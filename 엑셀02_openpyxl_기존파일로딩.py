import openpyxl as op 

#raw string natation: 가공하지 않은 문자열
wb = op.load_workbook(r"c:\work\test.xlsx")
#같은 폴더에 있는 파일
#wb = op.load_workbook("test.xlsx")

#새로운 시트 추가 
ws = wb.create_sheet("직원명부")

wb.save("test2.xlsx")

